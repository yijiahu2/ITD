from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROWS: list[dict[str, str]] = []


def format_metric_label(label: str) -> str:
    text = (label or "").strip()
    if not text:
        return ""
    match = re.match(r"^(.*?)[（(]([A-Za-z0-9_./%<>=:+@, \\-]+)[)）]$", text)
    if match:
        zh = match.group(1).strip()
        raw = match.group(2).strip()
        return f"{raw} / {zh}"
    if re.fullmatch(r"[A-Za-z0-9_./%<>=:+@, \\-]+", text):
        return text
    return f"- / {text}"


def add(
    category: str,
    level1_name: str,
    level1_formula_role: str,
    stage: str,
    level2_name: str = "",
    level2_formula_role: str = "",
    level3_name: str = "",
    level3_formula_role: str = "",
) -> None:
    ROWS.append(
        {
            "类别": category,
            "一级指标（中文名）": level1_name,
            "公式及作用_1": level1_formula_role,
            "二级指标（中文名）": level2_name,
            "公式及作用_2": level2_formula_role,
            "三级指标（中文名）": level3_name,
            "公式及作用_3": level3_formula_role,
            "使用环节": stage,
        }
    )


def build_rows() -> None:
    add(
        "在线输出质量",
        "在线输出质量总分（quality_score）",
        "quality_score = 0.35*|coverage_ratio-1| + 0.30*instance_leakage + 0.25*semantic_gap + 0.03*fragmentation_score + 0.02*merge_blob_score + 0.05*max(0, 0.60-instance_height_support_ratio)。作用：DOM-only 条件下给主模型结果、ROI 合并结果和最终结果提供统一在线质量误差分，分数越低越好。",
        "主模型首轮评估；ROI merged 结果在线评估；最终结果评估",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "由 data_processing.fusion.diagnostics 透传，用于评估实例结果与语义先验的一致性。",
        "在线质量打分；最终报告；几何诊断派生",
        "覆盖比（coverage_ratio）",
        "coverage_ratio = instance_union_area / semantic_area。作用：看实例覆盖相对语义先验是偏少还是偏多。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "在线质量解释；报告",
        "语义召回率（semantic_recall）",
        "semantic_recall = overlap_area / semantic_area。作用：衡量语义树冠区域被实例结果覆盖的比例。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "在线质量打分；报告",
        "实例泄漏率（instance_leakage）",
        "instance_leakage = max(instance_union_area-overlap_area,0) / instance_union_area。作用：衡量实例越界、误检外溢。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "在线质量打分；报告；派生 semantic_coverage_gap",
        "语义缺口率（semantic_gap）",
        "semantic_gap = max(semantic_area-overlap_area,0) / semantic_area。作用：衡量语义支持区域未被实例覆盖的漏检缺口。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "几何诊断派生；报告",
        "语义-实例 IoU（overlap_iou）",
        "overlap_iou = overlap_area / (semantic_area + instance_union_area - overlap_area)。作用：作为派生 semantic_instance_consistency 的底层依据。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "报告",
        "语义覆盖率（semantic_cover_ratio）",
        "semantic_cover_ratio = semantic_area / patch_area_m2。作用：给 patch 级语义覆盖背景。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "几何诊断派生；报告",
        "实例覆盖率（instance_cover_ratio）",
        "instance_cover_ratio = instance_union_area / patch_area_m2。作用：作为 pred_cover_ratio 的优先来源。",
    )
    add(
        "在线输出质量",
        "语义-实例一致性（原始事实块）",
        "同上。",
        "报告",
        "覆盖率差绝对值（cover_ratio_delta_abs）",
        "cover_ratio_delta_abs = |instance_cover_ratio - semantic_cover_ratio|。作用：解释实例覆盖和语义覆盖偏差。",
    )

    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "由 data_processing.fusion.diagnostics 透传，用于描述实例几何分布与形状健康度。",
        "在线质量解释；几何诊断派生；报告",
        "实例数（instance_count）",
        "instance_count = N_valid_instances。作用：原始实例规模统计。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告；几何解释",
        "面积和/union 比（sum_to_union_ratio）",
        "sum_to_union_ratio = sum_area_m2 / union_area_m2。作用：反映实例重叠和粘连程度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告",
        "平均面积（mean_area_m2）",
        "所有实例面积均值。作用：观察整体实例尺度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告",
        "面积中位数（median_area_m2）",
        "所有实例面积中位数。作用：减弱极端值影响。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告/分析",
        "面积分位数（p10_area_m2 / p90_area_m2）",
        "实例面积 10%/90% 分位值。作用：观察分布尾部。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告；几何解释",
        "等效冠幅均值/中位数",
        "equivalent_crown_width = 2*sqrt(area/pi)。作用：将面积转换到直观冠幅尺度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "在线质量打分（<4m²）；几何诊断派生；报告",
        "小碎片比例（<1/<2/<4/<6m²）",
        "small_fragment_ratio_lt_k = mean(area < k)。作用：描述碎片化程度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "shape_anomaly_ratio 派生",
        "极小冠幅比例（<1m/<2m）",
        "tiny_width_ratio_lt_k = mean(eq_width < k)。作用：辅助识别异常小目标。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "large_blob_ratio 派生",
        "大冠幅比例（>6m）",
        "large_width_ratio_gt_6m = mean(eq_width > 6)。作用：辅助识别异常大斑块/粘连。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "online quality fallback；shape_anomaly/large_blob/merge_blob 派生",
        "最大单块占比（max_instance_area_share）",
        "max_instance_area_share = max(area_i) / union_area_m2。作用：识别主导性大斑块。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告",
        "前5大单块占比（top5_instance_area_share）",
        "top5_instance_area_share = sum(top5 area) / union_area_m2。作用：观察大块集中度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "duplicate_overlap_ratio 派生；报告",
        "重叠对数（overlap_pair_count）",
        "实例对之间存在非零交叠的 pair 数。作用：反映重复预测/粘连。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "报告",
        "总重叠面积（overlap_area_total_m2）",
        "所有实例对交叠面积累加。作用：解释重叠严重度。",
    )
    add(
        "在线输出质量",
        "几何原始统计（geometry_plausibility）",
        "同上。",
        "edge_artifact_score 派生；报告",
        "边界接触数/率（edge_touch_count/edge_touch_ratio）",
        "edge_touch_ratio = edge_touch_count / instance_count。作用：识别边缘伪影。",
    )

    add(
        "在线输出质量",
        "高度支撑一致性（height_consistency）",
        "由 diagnostics 透传，仅在有 CHM 时生效。",
        "在线质量打分；报告",
        "高度支撑比例（instance_height_support_ratio）",
        "instance_height_support_ratio = support_pixels / instance_pixels，其中 support 定义为 CHM>1m 且有效像元。作用：衡量实例是否被高度证据支撑，也是 online quality 的附加惩罚项。",
    )
    add(
        "在线输出质量",
        "高度支撑一致性（height_consistency）",
        "同上。",
        "报告/分析",
        "高度均值/95分位/标准差",
        "height_mean / height_p95 / height_std。作用：描述实例内部高度分布。",
    )
    add(
        "在线输出质量",
        "高度支撑一致性（height_consistency）",
        "同上。",
        "报告/分析",
        "高度边缘强度（height_edge_strength）",
        "CHM 梯度在实例内的平均强度。作用：解释树高边缘清晰度。",
    )
    add(
        "在线输出质量",
        "高度支撑一致性（height_consistency）",
        "同上。",
        "报告/分析",
        "支撑面积（support_area）",
        "support_area = support_pixels * pixel_area。作用：表征高度证据空间规模。",
    )

    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "由 evaluation_analysis 对 raw geometry/semantic facts 做规则化诊断。",
        "在线质量解释；最终报告；决策 flags",
        "预测实例数（pred_instance_count）",
        "pred_instance_count = geometry.instance_count 或 valid_instance_count。作用：给评估层统一实例规模口径。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "在线质量解释；最终报告",
        "空输出标记（empty_output_flag）",
        "empty_output_flag = (pred_instance_count <= 0) OR (pred_cover_ratio < 0.005)。作用：快速识别几乎无有效输出。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "empty_output_flag；报告",
        "预测覆盖率（pred_cover_ratio）",
        "优先使用 instance_cover_ratio；若无则 union_area_m2 / patch_area_m2。作用：统一实例覆盖率口径。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "报告/诊断",
        "几何有效比例（valid_instance_ratio）",
        "valid_instance_ratio = valid_instance_count / raw_feature_count。作用：识别无效几何污染程度。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "fragmentation_score 派生；报告",
        "形状异常比例（shape_anomaly_ratio）",
        "shape_anomaly_ratio = clamp01(0.35*tiny_width_ratio_lt_1m + 0.35*small_fragment_ratio_lt_1m2 + 0.30*dominant_blob_norm)。作用：衡量极小碎片和异常主导块的综合异常。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "fragmentation_score；online quality",
        "小碎片比例（small_fragment_ratio）",
        "small_fragment_ratio = small_fragment_ratio_lt_4m2。作用：统一碎片化主口径。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "merge_blob_score；报告",
        "大斑块比例（large_blob_ratio）",
        "large_blob_ratio = clamp01(max(large_width_ratio_gt_6m, max_instance_area_share))。作用：识别大块粘连。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "报告/诊断",
        "重复重叠比例（duplicate_overlap_ratio）",
        "duplicate_overlap_ratio = overlap_pair_count / pred_instance_count。作用：衡量重复预测/实例交叠。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "报告/诊断",
        "边缘伪影分数（edge_artifact_score）",
        "edge_artifact_score = edge_touch_ratio。作用：识别边缘接触伪影。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "online quality；报告",
        "碎片化分数（fragmentation_score）",
        "fragmentation_score = clamp01(0.65*small_fragment_ratio + 0.35*shape_anomaly_ratio)。作用：给在线质量打分和几何诊断提供统一碎片化分数。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "online quality；报告",
        "粘连分数（merge_blob_score）",
        "merge_blob_score = clamp01(0.60*large_blob_ratio + 0.40*(1-semantic_instance_consistency))。作用：给在线质量打分和粘连诊断提供统一分数。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "merge_blob_score；冲突判定；报告",
        "语义-实例一致性（semantic_instance_consistency）",
        "semantic_instance_consistency = overlap_iou。作用：把原始 overlap_iou 规范成评估层主指标。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "冲突判定；报告",
        "语义缺口（semantic_coverage_gap）",
        "semantic_coverage_gap = semantic_gap。作用：统一语义缺口主口径。",
    )
    add(
        "在线输出质量",
        "几何诊断派生（geometry_diagnostics）",
        "同上。",
        "decision_flags.need_manual_review_flag；报告",
        "语义-实例冲突标记（semantic_instance_conflict_flag）",
        "semantic_instance_conflict_flag = (overlap_iou < 0.50) OR (semantic_gap > 0.30)。作用：触发人工复核建议。",
    )

    add(
        "参考质量评估",
        "参考质量分（quality_score / reference score）",
        "score = Σ(weight_i * normalized_metric_i)，默认权重为 tree_count 0.30、crown 0.40、closure 0.20、density 0.10；当 tree_count_error_ratio <= 0.12 时切到 boundary_priority 权重（0.15/0.55/0.20/0.10）。作用：主模型、专家模型、ROI 轮次和最终 reference 结果的统一误差分，分数越低越好。",
        "主模型首轮（有 reference 时）；专家模型比较；ROI 轮次；最终 reference 评估",
    )
    add(
        "参考质量评估",
        "树木数量一致性",
        "以树木数量相对误差衡量实例数与参考约束的一致性。",
        "reference score；ROI 触发；flow decision",
        "树木数量误差率（tree_count_error_ratio）",
        "tree_count_error_ratio = |pred_tree_count - expected_tree_count| / max(expected_tree_count, 1e-6)。作用：参考评分、ROI 触发、主模型和专家模型对比。",
    )
    add(
        "参考质量评估",
        "树木数量一致性",
        "同上。",
        "details summary；报告",
        "树木数量误差绝对值（tree_count_error_abs）",
        "tree_count_error_abs = |pred_tree_count - expected_tree_count|。作用：报告解释、细节定位。",
    )
    add(
        "参考质量评估",
        "冠幅边界一致性",
        "以平均冠幅相对误差衡量边界恢复程度。",
        "reference score；ROI 触发；flow decision",
        "冠幅误差率（mean_crown_width_error_ratio）",
        "mean_crown_width_error_ratio = |pred_mean_crown_width - expected_mean_crown_width| / max(expected_mean_crown_width, 1e-6)。作用：参考评分、ROI 触发、专家模型比较。",
    )
    add(
        "参考质量评估",
        "冠幅边界一致性",
        "同上。",
        "details summary；报告",
        "冠幅误差绝对值（mean_crown_width_error_abs）",
        "mean_crown_width_error_abs = |pred_mean_crown_width - expected_mean_crown_width|。作用：报告解释、问题单元排序。",
    )
    add(
        "参考质量评估",
        "郁闭度一致性",
        "以预测覆盖率和参考郁闭度的偏差衡量覆盖恢复程度。",
        "reference score；ROI 触发；flow decision",
        "郁闭度误差绝对值（closure_error_abs）",
        "closure_error_abs = |pred_cover_ratio - expected_closure|。作用：参考评分、ROI 触发、专家模型比较。",
    )
    add(
        "参考质量评估",
        "林分密度一致性",
        "以预测密度与参考密度偏差衡量数量尺度一致性。",
        "reference score 输入；flow decision；报告",
        "密度误差绝对值（density_error_abs）",
        "density_error_abs = |pred_density_trees_per_ha - expected_density|。作用：报告、专家模型比较、decision_flags 归一化。",
    )
    add(
        "参考质量评估",
        "林分密度一致性",
        "同上。",
        "reference score",
        "密度误差率（density_error_ratio）",
        "density_error_ratio = density_error_abs / expected_density；若 expected_density 缺失，则退化为 density_error_abs / 1000。作用：进入参考质量分加权。",
    )
    add(
        "参考质量评估",
        "Patch/参考背景统计",
        "辅助解释 patch 与参考约束的匹配背景。",
        "reference 评估；报告；细节排序",
        "预测树木数量/平均冠幅/覆盖率/密度",
        "pred_tree_count、pred_mean_crown_width、pred_cover_ratio、pred_density_trees_per_ha。作用：作为参考误差计算的预测端输入。",
    )
    add(
        "参考质量评估",
        "Patch/参考背景统计",
        "同上。",
        "reference 评估；报告；细节排序",
        "期望树木数量/平均冠幅/郁闭度/密度",
        "expected_tree_count、expected_mean_crown_width、expected_closure、expected_density。作用：作为参考误差计算的参考端输入。",
    )
    add(
        "参考质量评估",
        "地形分层误差摘要",
        "按地形类别聚合误差均值，用于解释误差分布。",
        "reference 结果解释；报告；memory/finetune",
        "terrain_stratified_error_summary",
        "按 landform_type / slope_class / aspect_class / slope_position_class 分组，统计 tree/crown/closure/density 误差均值。",
    )
    add(
        "参考质量评估",
        "问题参考单元排序",
        "detail_ranker 按 reference score 从高到低排序问题单元。",
        "ROI fallback；报告；memory；finetune pool",
        "top_k_reference_units",
        "对 details.csv 中每个参考单元计算 error_score，并输出 top-k 最差单元。作用：ROI 候选回退、报告解释、memory/finetune 失败样本抽取。",
        "error_score",
        "error_score = build_reference_score_breakdown(单元级 metrics).score。作用：排序问题单元。",
    )

    add(
        "标准真值 Benchmark",
        "Benchmark 综合质量",
        "基于 GT/COCO 树冠矢量的标准实例分割质量评估。",
        "最终 benchmark 评估；benchmark gain；decision_flags overall_score",
        "AP50（ap50）",
        "PR 曲线在 IoU>=0.50 下积分。作用：最核心的实例分割质量指标。",
    )
    add(
        "标准真值 Benchmark",
        "Benchmark 综合质量",
        "同上。",
        "最终 benchmark 评估；benchmark gain；decision_flags overall_score",
        "AP75（ap75）",
        "PR 曲线在 IoU>=0.75 下积分。作用：更严格的边界质量指标。",
    )
    add(
        "标准真值 Benchmark",
        "Benchmark 综合质量",
        "同上。",
        "benchmark 评估；flow decision；报告",
        "Precision / Recall",
        "precision = TP/(TP+FP)；recall = TP/GT。作用：检测级正确率和召回率。",
    )
    add(
        "标准真值 Benchmark",
        "Benchmark 综合质量",
        "同上。",
        "benchmark 评估；decision_flags overall_score；报告",
        "F1@0.50（f1_score50）",
        "f1_score50 = 2*precision*recall/(precision+recall)。作用：单值概括 precision/recall 折中。",
    )
    add(
        "标准真值 Benchmark",
        "Benchmark 综合质量",
        "同上。",
        "benchmark 评估；报告",
        "平均匹配 IoU（mean_iou_matched）",
        "所有 IoU>=0.50 的 TP 匹配实例 best_iou 的均值。作用：解释匹配实例边界贴合程度。",
    )
    add(
        "标准真值 Benchmark",
        "匹配计数",
        "记录不同 IoU 阈值下的 TP/FP/FN。",
        "benchmark 评估；error decomposition",
        "TP50/FP50/FN50",
        "由贪心匹配产生。作用：支撑 precision/recall 和错误分解。",
    )
    add(
        "标准真值 Benchmark",
        "匹配计数",
        "同上。",
        "benchmark 评估；报告",
        "TP75/FP75/FN75",
        "在 IoU>=0.75 下统计。作用：更严格的实例匹配计数。",
    )
    add(
        "标准真值 Benchmark",
        "面积回归质量",
        "仅对 TP 匹配实例的面积做回归误差分析。",
        "benchmark 评估；benchmark gain；报告",
        "MAE / RMSE / RMSE% / R2",
        "基于 matched_gt_area 与 pred_area 的差值计算。作用：评估树冠面积拟合质量。",
        "匹配树冠数量（crown_area_iou_0_50.num_matched_crowns）",
        "TP 行数。作用：解释面积回归指标样本量。",
    )

    add(
        "错误分解",
        "过/欠分割与检错分解",
        "将 benchmark 匹配结果进一步分解为过分割、欠分割、漏检、误检。",
        "benchmark 评估；manual review",
        "欠分割分数（under_segmentation_score）",
        "under_segmentation_score = pred_multi_gt / pred_count，其中 pred_multi_gt 表示一个 pred 对应多个 GT。作用：识别合并过度。",
    )
    add(
        "错误分解",
        "过/欠分割与检错分解",
        "同上。",
        "benchmark 评估；manual review",
        "过分割分数（over_segmentation_score）",
        "over_segmentation_score = gt_multi_pred / gt_count，其中 gt_multi_pred 表示一个 GT 被多个 pred 覆盖。作用：识别切分过细。",
    )
    add(
        "错误分解",
        "过/欠分割与检错分解",
        "同上。",
        "benchmark 评估；manual review",
        "漏检分数（miss_detection_score）",
        "miss_detection_score = FN50 / GT_count。作用：归一化漏检程度。",
    )
    add(
        "错误分解",
        "过/欠分割与检错分解",
        "同上。",
        "benchmark 评估；manual review",
        "误检分数（false_detection_score）",
        "false_detection_score = FP50 / pred_count。作用：归一化误检程度。",
    )
    add(
        "错误分解",
        "错误置信度",
        "用最严重错误和次严重错误差距估计当前主导错误模式的置信度。",
        "benchmark 评估；decision_flags",
        "failure_confidence",
        "failure_confidence = clamp01(top + 0.35*(top-second))。作用：触发 need_manual_review_flag。",
    )

    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "基于参考误差阈值、候选 ROI 数量和轮次控制是否继续局部细化。",
        "ROI 首轮评估；每轮 ROI 继续判定；专家模型比较",
        "当前参考分（current_score）",
        "current_score = build_reference_score_breakdown(metrics).score。作用：当前轮次参考误差分，越低越好。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "每轮 ROI 继续判定；专家模型比较",
        "前一轮参考分（previous_score）",
        "上一轮或主模型基线的参考误差分。作用：计算 improvement。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "ROI 继续判定；专家模型接收",
        "改进量（improvement）",
        "improvement = previous_score - current_score。作用：判断 ROI 是否有增益。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "ROI 触发；decision_flags.need_local_refine_flag",
        "候选 ROI 数量（candidate_roi_count）",
        "candidate_roi_count = len(candidate_rois)。作用：判定是否有继续细化价值。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "ROI 首轮；每轮 ROI 决策",
        "触发指标列表（trigger_metrics）",
        "当 tree_count_error_ratio / mean_crown_width_error_ratio / closure_error_abs 超阈值，或存在 problem_roi_cases 时加入列表。作用：说明为什么继续细化。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "ROI 决策",
        "阈值配置（metric_thresholds）",
        "tree_count_error_ratio 默认 0.18/0.10；mean_crown_width_error_ratio 默认 0.12/0.18；closure_error_abs 默认 0.06/0.10，且 round>0 时进一步收紧。作用：控制 ROI 触发敏感度。",
    )
    add(
        "ROI 细化决策",
        "ROI 继续/停止判定",
        "同上。",
        "ROI 决策；planning/调度",
        "heuristic_continue / continue_refinement",
        "heuristic_continue = enabled AND round_idx<max_rounds AND 有触发指标 AND candidate_rois>=1 AND (无 improvement 或 improvement>=-epsilon)。作用：给 orchestration 明确的继续/停止信号。",
    )
    add(
        "ROI 细化决策",
        "问题参考单元回退候选",
        "当上游没有 signal-driven ROI candidate 时，回退使用 detail_ranker 的最差参考单元。",
        "ROI 首轮和各轮评估",
        "candidate_source",
        "candidate_source = precomputed 或 inventory_detail_fallback。作用：解释候选来源。",
    )

    add(
        "决策 Flags",
        "综合总分与流程建议",
        "将在线质量、reference/benchmark 质量、错误分解与 ROI 信号转成统一流程建议。",
        "主模型；最终结果；专家模型/微调比较",
        "overall_score",
        "benchmark 模式：0.40*ap50 + 0.35*ap75 + 0.25*f1_score50；reference/online 模式：reference_score 与 (1-online_quality_score) 线性组合。作用：统一总览质量分，越高越好。",
    )
    add(
        "决策 Flags",
        "综合总分与流程建议",
        "同上。",
        "主模型；最终结果",
        "quality_pass_flag",
        "quality_pass_flag = overall_score >= pass_threshold。作用：判断结果是否达到可接受线。",
    )
    add(
        "决策 Flags",
        "综合总分与流程建议",
        "同上。",
        "主模型；ROI",
        "need_local_refine_flag",
        "need_local_refine_flag = continue_refinement OR len(candidate_rois)>0 OR candidate_roi_count>0。作用：建议是否进入/继续 ROI。",
    )
    add(
        "决策 Flags",
        "综合总分与流程建议",
        "同上。",
        "主模型；最终结果",
        "need_param_search_flag",
        "need_param_search_flag = overall_score < param_search_threshold AND not quality_pass_flag。作用：建议是否重新参数搜索。",
    )
    add(
        "决策 Flags",
        "综合总分与流程建议",
        "同上。",
        "主模型；最终结果",
        "need_finetune_flag",
        "need_finetune_flag = overall_score < finetune_threshold AND not quality_pass_flag AND not regression_flag。作用：建议是否进入微调池。",
    )
    add(
        "决策 Flags",
        "综合总分与流程建议",
        "同上。",
        "最终结果；benchmark 分析",
        "need_manual_review_flag",
        "need_manual_review_flag = semantic_instance_conflict_flag OR (failure_confidence < manual_review_confidence_threshold)。作用：建议人工复核。",
    )
    add(
        "决策 Flags",
        "结果比较标记",
        "用于专家模型接收和微调效果比较。",
        "专家模型比较；微调前后比较",
        "accepted_improvement_flag",
        "accepted_improvement_flag = overall_score >= previous_overall_score + accepted_gain_threshold AND not regression_flag；微调比较时若 benchmark delta 或均值增益足够也会置 true。作用：自动接受更优结果。",
    )
    add(
        "决策 Flags",
        "结果比较标记",
        "同上。",
        "专家模型比较；微调前后比较",
        "regression_flag",
        "regression_flag = overall_score < previous_overall_score - regression_threshold；微调比较时若 benchmark AP50 下降或负向增益占优也会置 true。作用：阻止退化结果被接受。",
    )

    add(
        "微调效果评估",
        "微调前后效果对比",
        "对 before/after details.csv 和可选 benchmark 结果做收益闭环评估。",
        "微调效果评估；flow decision",
        "mean_gain_tree_count / crown / closure / density",
        "gain_x = before_error_abs - after_error_abs；均值为所有参考单元的平均改善量。作用：量化各误差项平均提升幅度。",
    )
    add(
        "微调效果评估",
        "微调前后效果对比",
        "同上。",
        "微调效果评估；flow decision",
        "num_tree_improved / num_crown_improved / num_closure_improved / num_density_improved",
        "统计 gain_x > 0 的参考单元数量。作用：量化提升覆盖面。",
    )
    add(
        "微调效果评估",
        "微调前后效果对比",
        "同上。",
        "微调效果评估；报告/分析",
        "stratified_gain",
        "按 landform_type / slope_class / aspect_class / slope_position_class 分组统计各 gain 的均值。作用：分析微调在不同场景下的收益差异。",
    )
    add(
        "微调效果评估",
        "微调前后效果对比",
        "同上。",
        "微调效果评估",
        "benchmark_gain.delta",
        "对 before/after benchmark 的 precision/recall/ap50/ap75/mae/rmse/rmse_percent/r2 做差；对误差项使用 before-after。作用：形成真值条件下的提升闭环。",
    )

    add(
        "流程决策输出",
        "flow_decision 结构化输出",
        "把不同阶段真正服务决策的核心指标和证据指标结构化输出给 orchestration、report、retrospective。",
        "主模型评估；ROI；专家模型；最终结果；微调效果",
        "decision_stage / decision_question / core_metrics / evidence_metrics / decision",
        "不是新公式，而是评估结果的结构化组织层。作用：保证不同阶段都能按统一格式消费指标。",
    )


def build_workbook(output_path: Path) -> None:
    build_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "指标总表"

    headers = [
        "类别",
        "一级指标（原始指标名称/中文名）",
        "公式及作用",
        "二级指标（原始指标名称/中文名）",
        "公式及作用",
        "三级指标（原始指标名称/中文名）",
        "公式及作用",
        "使用环节",
    ]
    ws.append(headers)

    for row in ROWS:
        ws.append(
            [
                row["类别"],
                format_metric_label(row["一级指标（中文名）"]),
                row["公式及作用_1"],
                format_metric_label(row["二级指标（中文名）"]),
                row["公式及作用_2"],
                format_metric_label(row["三级指标（中文名）"]),
                row["公式及作用_3"],
                row["使用环节"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = {1: 16, 2: 28, 3: 56, 4: 28, 5: 56, 6: 28, 7: 56, 8: 26}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    note = wb.create_sheet("说明")
    note.append(["字段", "说明"])
    notes = [
        ["范围", "仅基于当前仓库代码真实实现的 evaluation_analysis 及其直接消费的 diagnostics / decision flags / benchmark / finetune 指标整理。"],
        ["口径", "表中既包含透传事实指标，也包含 evaluation_analysis 派生诊断指标、综合评分和流程决策 flags。"],
        ["分数方向", "quality_score / reference score 越低越好；overall_score 越高越好。"],
        ["生成时间", "2026-05-07 Asia/Shanghai"],
    ]
    for item in notes:
        note.append(item)
    for cell in note[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in note.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    note.column_dimensions["A"].width = 18
    note.column_dimensions["B"].width = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    output_path = Path("/home/xth/forest_agent_project/docs/evaluation_analysis_完整指标表_20260507.xlsx")
    build_workbook(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
